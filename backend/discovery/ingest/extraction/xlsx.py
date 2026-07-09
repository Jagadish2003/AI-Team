"""
R18-A1 / T2 — Excel (.xlsx) text extraction handler.

Flattens each worksheet's cell content to text with ``openpyxl`` (read-only,
values-only), so the knowledge inside a spreadsheet becomes searchable prose. Each
sheet is prefixed with a ``# Sheet: <name>`` header and its rows are rendered as
tab-separated cell values (empty trailing cells trimmed). An encrypted workbook
(OLE compound file, not a ZIP) is a loud
:data:`~discovery.ingest.extraction.ENCRYPTED` skip (AC3); a corrupt/unparseable
one raises :class:`~discovery.ingest.extraction.ExtractionError` so the ingestor
isolates it per-file (AC5). The legacy binary ``.xls`` format is out of phase-one
scope.
"""
from __future__ import annotations

import io
import logging

from . import (
    ENCRYPTED,
    NO_HANDLER,
    ExtractedText,
    ExtractionError,
    ExtractionSkipped,
    ExtractionOutcome,
    register_handler,
)
from ._office_common import looks_encrypted_ooxml

logger = logging.getLogger(__name__)

#: A cap on the cells rendered per workbook, so one giant sheet cannot dominate a
#: run. This is a defensive bound on the extraction WORK; the per-file size cap and
#: per-run extraction budget are the ingestor's concern (T4).
_MAX_CELLS = 200_000


def extract_xlsx(raw: bytes, fmt: str) -> ExtractionOutcome:
    """Extract sheet/cell text from a .xlsx, or record a loud skip."""
    if looks_encrypted_ooxml(raw):
        return ExtractionSkipped(ENCRYPTED, "Excel workbook is password-protected")

    try:
        import openpyxl
    except ImportError:  # pragma: no cover - openpyxl ships in requirements
        return ExtractionSkipped(NO_HANDLER, "openpyxl is not installed")

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True
        )
    except Exception as exc:  # noqa: BLE001 — a parse failure is a corrupt file (AC5)
        raise ExtractionError(f"could not parse .xlsx: {type(exc).__name__}") from exc

    try:
        sheet_names = list(workbook.sheetnames)
        parts = []
        cells = 0
        for name in sheet_names:
            sheet = workbook[name]
            parts.append(f"# Sheet: {name}")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                # Trim trailing empties so blank columns do not bloat the text.
                while values and values[-1] == "":
                    values.pop()
                cells += len(values)
                if values:
                    parts.append("\t".join(values))
                if cells >= _MAX_CELLS:
                    parts.append("… (truncated: cell budget reached)")
                    break
            if cells >= _MAX_CELLS:
                break
    finally:
        workbook.close()

    content = "\n".join(parts).strip()
    return ExtractedText(
        content=content,
        chunk_content_type="prose",
        structure_hints={
            "format": "xlsx",
            "sheets": sheet_names,
            "sheet_count": len(sheet_names),
            "chars": len(content),
        },
    )


register_handler("xlsx", extract_xlsx)
